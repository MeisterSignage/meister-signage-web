import re, glob, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

NEWS = "content/news"
SCHED = {
    "spark4-vielseitiger-bestseller": "2026-06-09",
    "digital-signage-events-tipps": "2026-06-11",
    "digital-signage-kosten-schweiz": "2026-06-16",
    "spark5-neues-modell": "2026-06-18",
    "digital-signage-mieten-oder-kaufen": "2026-06-23",
    "digital-signage-gastronomie-2026": "2026-06-25",
    "was-ist-digital-signage": "2026-06-30",
    "spark-sortiment-vergleich": "2026-07-02",
    "digitales-menueboard-vorteile": "2026-07-07",
    "digital-signage-retail": "2026-07-14",
    "digital-signage-hotellerie": "2026-07-21",
    "spark3-kompaktes-display": "2026-07-28",
    "sparkq-quadratisches-display": "2026-08-04",
    "digital-signage-kmu": "2026-08-11",
}

titles = {}
for f in glob.glob(NEWS + "/*.md"):
    t = open(f, encoding="utf-8").read()
    s = re.search(r'^slug:\s*"?(.+?)"?\s*$', t, re.M)
    ti = re.search(r'^title:\s*"?(.+?)"?\s*$', t, re.M)
    slug = s.group(1) if s else os.path.basename(f)[:-3]
    titles[slug] = ti.group(1) if ti else slug

rows = []
for slug, ds in SCHED.items():
    d = datetime.date.fromisoformat(ds)
    name = titles.get(slug, slug)
    rows.append((name, "LinkedIn", d, "11:00"))
    rows.append((name, "Instagram", d, "12:00"))
rows.sort(key=lambda r: (r[2], r[3]))

wb = Workbook(); ws = wb.active; ws.title = "Posting-Plan"
ws.append(["Postname", "Kanal", "Datum", "Zeit"])
for r in rows:
    ws.append([r[0], r[1], r[2], r[3]])

hdr_fill = PatternFill("solid", start_color="1A2744")
for c in ws[1]:
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="left", vertical="center")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        c.font = Font(name="Arial", size=11)
    row[2].number_format = "DD.MM.YYYY"
    ch = row[1].value
    row[1].font = Font(name="Arial", bold=True, color="0A66C2" if ch == "LinkedIn" else "C13584")

ws.column_dimensions['A'].width = 54
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 10
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:D1"
wb.save("docs/social-schedule.xlsx")
print("rows:", len(rows))
